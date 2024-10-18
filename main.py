import asyncio
import datetime
import os
import sqlite3
import time
from itertools import count
from typing import Union, List
from attr import dataclass

import aiogram
from aiogram import exceptions
from aiogram.types import Message, CallbackQuery, FSInputFile, \
    InlineKeyboardMarkup as Markup, InlineKeyboardButton as Button
from dotenv import load_dotenv

import texts


class DotEnvData:
    BASE_CHAT_ID: int
    REPORT_CHAT_ID: int
    BOT_TOKEN: str
    BASE_CHAT_LINK: str
    ADMIN_ID: int

    def __init__(self):
        environ = os.environ
        if environ.get('BASE_CHAT_ID') is None:
            load_dotenv('.env')
            environ = os.environ
        if environ.get('BASE_CHAT_ID') is None:
            raise Exception('Необходимо в .env указать основной чат (канал, группа)')

        self.BASE_CHAT_ID = int(environ.get('BASE_CHAT_ID') or 0)
        self.REPORT_CHAT_ID = int(environ.get('REPORT_CHAT_ID') or 0)
        self.ADMIN_ID = int(environ.get('ADMIN_ID'))
        self.BOT_TOKEN = environ.get('BOT_TOKEN') or ''
        self.BASE_CHAT_LINK = environ.get('BASE_CHAT_LINK') or ''


@dataclass
class DbUser:
    id: int
    user_id: int
    username: Union[str, None]
    first_name: Union[str, None]
    subscribe: int


class DataBase:
    def __init__(self):
        self.conn = sqlite3.connect('database.db')
        cursor = self.conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS user (
                id INTEGER PRIMARY KEY,
                user_id INTEGER NOT NULL,
                username TEXT DEFAULT NULL,
                first_name TEXT DEFAULT NULL,
                subscribe INTEGER DEFAULT 0
            );
        """)
        self.conn.commit()

    async def execute(self, query: str, args: dict) -> None:
        cursor = self.conn.cursor()
        cursor.execute(query, args)
        self.conn.commit()
        cursor.close()

    async def fetch(self, query: str, args: dict = None) -> Union[List[DbUser], None]:
        if not args:
            args = {}
        cursor = self.conn.cursor()
        cursor.execute(query, args)
        result = cursor.fetchall()
        cursor.close()

        if result:
            ret_list = []
            for i in result:
                ret_list.append(DbUser(
                    id=i[0],
                    user_id=i[1],
                    username=i[2],
                    first_name=i[3],
                    subscribe=i[4]
                ))
            return ret_list
        else:
            return None

    async def fetchrow(self, query: str, args: dict = None) -> Union[DbUser, None]:
        if not args:
            args = {}
        cursor = self.conn.cursor()
        cursor.execute(query, args)
        result = cursor.fetchone()
        cursor.close()
        if result:
            return DbUser(
                id=result[0],
                user_id=result[1],
                username=result[2],
                first_name=result[3],
                subscribe=result[4]
            )
        else:
            return None


db = DataBase()

EnvData = DotEnvData()

bot = aiogram.Bot(token=EnvData.BOT_TOKEN)
dispatcher = aiogram.Dispatcher()

allowed_updates = ['message', 'callback_query']


# noinspection PyBroadException
@dispatcher.message()
async def tg_message(message: Message):
    try:
        if message.chat.type != 'private':
            return
        if not message.text:
            return

        user_id = message.from_user.id

        db_user = await db.fetchrow(f"""
            SELECT * FROM user
            WHERE user_id = :user_id;
        """, {'user_id': user_id})

        username = message.from_user.username
        first_name = message.from_user.first_name

        if not db_user:
            await db.execute(f"""
                INSERT INTO user (user_id, username, first_name, subscribe)
                VALUES (:user_id, :username, :first_name, 0);
            """, {'user_id': user_id, 'username': username, 'first_name': first_name})
            mention = f'@{username}' if username else f'id={user_id}'
            await bot.send_message(
                chat_id=EnvData.REPORT_CHAT_ID,
                text=f'@{mention} нажал /start'
            )

        text_low = message.text.lower().strip()
        split_n_1 = message.text.split('\n', 1)

        if text_low == '/start':
            await bot.send_photo(
                chat_id=user_id,
                caption=texts.first_text,
                photo=FSInputFile('first_photo.JPG'),
                reply_markup=Markup(inline_keyboard=[
                    [Button(text='ХОЧУ ПОПАСТЬ', callback_data='want_in')],
                    [Button(text='ЧТО ВНУТРИ СООБЩЕСТВА', callback_data='what_in')]
                ])
            )
        elif text_low == '/table' and user_id == EnvData.ADMIN_ID:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Выгрузка базы данных'

            res = await db.fetch("""
                SELECT * FROM user;
            """)

            ws.column_dimensions['A'].width = 3
            ws.column_dimensions['B'].width = 13
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 12

            ws.cell(row=1, column=1, value='ID')
            ws.cell(row=1, column=2, value='USER_ID')
            ws.cell(row=1, column=3, value='USERNAME')
            ws.cell(row=1, column=4, value='NAME')
            ws.cell(row=1, column=5, value='SUBSCRIBE')

            row = 2

            for i in res:
                ws.cell(row=row, column=1, value=i.id)
                ws.cell(row=row, column=2, value=i.user_id)
                ws.cell(row=row, column=3, value=i.username or ' ')
                ws.cell(row=row, column=4, value=i.first_name or ' ')
                if i.subscribe == 1:
                    ws.cell(row=row, column=5, value='✅')
                elif i.subscribe == 0:
                    ws.cell(row=row, column=5, value='❌')
                row += 1

            today = datetime.datetime.today().strftime(format='%Y.%m.%d')

            file_name = f'Выгрузка {today}.xlsx'

            wb.save(file_name)

            await bot.send_document(
                chat_id=user_id,
                document=FSInputFile(file_name)
            )

            os.remove(file_name)
        elif split_n_1[0].lower() == 'рассылка' and user_id == EnvData.ADMIN_ID:
            res = await db.fetch("""
                SELECT * FROM user
                WHERE subscribe = 1;
            """)

            count_users = 0

            for i in res:
                try:
                    await bot.send_message(
                        chat_id=i.user_id,
                        text=split_n_1[1]
                    )
                    count_users += 1
                except:
                    pass

            await bot.send_message(
                chat_id=message.from_user.id,
                text=f'Рассылка успешно разослана количеству пользователей: {str(count_users)}'
            )


    except Exception as err:
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")


# noinspection PyBroadException
@dispatcher.callback_query()
async def tg_callback(callback: CallbackQuery):
    try:
        if callback.message.chat.type != 'private':
            return

        user_id = callback.from_user.id

        cdata = callback.data

        link = EnvData.BASE_CHAT_LINK
        if not link:
            link = (await bot.get_chat(EnvData.BASE_CHAT_ID)).invite_link
        if not link:
            link = (await bot.create_chat_invite_link(EnvData.BASE_CHAT_ID)).invite_link

        if cdata == 'want_in':
            await bot.send_message(
                chat_id=user_id,
                text=texts.second_text,
                reply_markup=Markup(inline_keyboard=[
                    [Button(text='Подписаться', url=link)],
                    [Button(text='Я подписался', callback_data='i_subscribe')]
                ])
            )
            try:
                await bot.delete_message(
                    chat_id=user_id,
                    message_id=callback.message.message_id
                )
            except:
                pass
        elif cdata == 'what_in':
            await bot.send_message(
                chat_id=user_id,
                text=texts.third_text,
                reply_markup=Markup(inline_keyboard=[
                    [Button(text='ХОЧУ ПОПАСТЬ', callback_data='want_in')]
                ])
            )
            try:
                await bot.delete_message(
                    chat_id=user_id,
                    message_id=callback.message.message_id
                )
            except Exception:
                pass
        elif cdata == 'i_subscribe':
            try:
                status = (await bot.get_chat_member(
                    chat_id=EnvData.BASE_CHAT_ID,
                    user_id=user_id
                )).status
                if status == 'left':
                    subscribed = False
                else:
                    subscribed = True
            except aiogram.exceptions.TelegramBadRequest:
                subscribed = False

            if subscribed is True:
                await bot.send_message(
                    chat_id=user_id,
                    text=texts.subscribe
                )
                await db.execute("""
                    UPDATE user
                    SET subscribe = 1
                    WHERE user_id = :user_id;
                """, {'user_id': user_id})

                mention = f'@{callback.from_user.username}' if callback.from_user.username else f'id={user_id}'

                await bot.send_message(
                    chat_id=EnvData.REPORT_CHAT_ID,
                    text=f'{mention})" подписался на канал и подал заявку'
                )

            else:
                await bot.send_message(
                    chat_id=user_id,
                    text=texts.not_subscribe,
                    reply_markup=Markup(inline_keyboard=[
                        [Button(text='Подписаться', url=link)],
                        [Button(text='Я подписался', callback_data='i_subscribe')]
                    ])
                )
            try:
                await bot.delete_message(
                    chat_id=user_id,
                    message_id=callback.message.message_id
                )
            except Exception:
                pass
    except Exception as err:
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")


async def main():
    while True:
        try:
            await dispatcher.start_polling(bot, polling_timeout=300, allowed_updates=allowed_updates)
        except Exception as err:
            print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
            time.sleep(3)


if __name__ == '__main__':
    asyncio.run(main())


"""
Криптовалюта на простом языке

Телеграм ATRUMS — https://t.me/+Gsb5CVCNsytjN2Y6
YouTube — https://www.youtube.com/@ATRUMS
"""